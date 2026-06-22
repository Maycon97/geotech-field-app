"""Build the embedded GeoView operational dataset from corporate workbooks."""

from __future__ import annotations

import json
import zipfile
from datetime import date, datetime
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook


CORPORATE_ROOT = Path(r"C:\Users\maycon.nascimento\ITAMINAS\SPLO - General")
DOWNLOADS_ROOT = Path(r"C:\Users\maycon.nascimento\Downloads")
OUTPUT_FILE = Path(__file__).resolve().parents[1] / "data" / "geoview-operational.js"


def iso_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def find_file(name: str, required_fragment: str) -> Path:
    matches = [
        path
        for path in CORPORATE_ROOT.rglob(name)
        if required_fragment.lower() in str(path).lower()
    ]
    if not matches:
        raise FileNotFoundError(f"Nao foi encontrado {name} em {required_fragment}.")
    return max(matches, key=lambda path: path.stat().st_mtime)


def read_pile_metrics(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["ITAMINAS"]
    rows = []
    for cells in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in cells):
            continue
        rows.append(
            {
                "pile": cells[0],
                "owner": cells[1],
                "analysisType": cells[2],
                "material": cells[3],
                "analysisDate": iso_value(cells[4]),
                "surveyDate": iso_value(cells[5]),
                "geometry": cells[6],
                "ponding": cells[7],
                "slopeConformity": cells[8],
                "minePlanSafetyFactor": cells[9],
                "safetyFactor": cells[10],
                "sectionUrl": cells[11],
            }
        )

    coordinate_sheet = workbook["ESTRUTURAS"]
    structures = []
    for cells in coordinate_sheet.iter_rows(min_row=2, values_only=True):
        if not cells[0]:
            continue
        structures.append(
            {
                "name": cells[0],
                "latitude": cells[1],
                "longitude": cells[2],
                "imageUrl": cells[3],
            }
        )
    return rows, structures


def read_rainfall(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["PLUVIOMETRIA"]
    rows = []
    for cells in worksheet.iter_rows(min_row=2, values_only=True):
        reading_date, _, responsible, location, millimeters = cells[:5]
        if not reading_date or not location or millimeters in (None, ""):
            continue
        rows.append(
            {
                "date": iso_value(reading_date),
                "location": str(location).strip(),
                "millimeters": float(millimeters),
                "responsible": responsible,
            }
        )
    return rows


def parse_coordinates(text: str):
    coordinates = []
    for chunk in str(text or "").strip().split():
        values = chunk.split(",")
        if len(values) < 2:
            continue
        try:
            longitude = float(values[0])
            latitude = float(values[1])
            altitude = float(values[2]) if len(values) > 2 else None
        except ValueError:
            continue
        coordinates.append(
            {
                "longitude": longitude,
                "latitude": latitude,
                "altitude": altitude,
            }
        )
    return coordinates


def read_kmz_layer(path: Path):
    with zipfile.ZipFile(path) as archive:
        kml_name = next(
            (name for name in archive.namelist() if name.lower().endswith(".kml")),
            None,
        )
        if not kml_name:
            return None
        root = ElementTree.fromstring(archive.read(kml_name))

    features = []
    for placemark in [node for node in root.iter() if node.tag.endswith("Placemark")]:
        name = next(
            (
                (node.text or "").strip()
                for node in placemark.iter()
                if node.tag.endswith("name")
            ),
            "Ponto Google Earth",
        )
        for geometry_type, feature_type in (
            ("Point", "point"),
            ("LineString", "line"),
            ("Polygon", "polygon"),
        ):
            for geometry in [
                node for node in placemark.iter() if node.tag.endswith(geometry_type)
            ]:
                coordinate_text = next(
                    (
                        node.text
                        for node in geometry.iter()
                        if node.tag.endswith("coordinates")
                    ),
                    "",
                )
                coordinates = parse_coordinates(coordinate_text)
                if coordinates:
                    features.append(
                        {
                            "type": feature_type,
                            "name": name,
                            "coordinates": coordinates,
                        }
                    )
    if not features:
        return None
    return {
        "fileName": path.name,
        "name": path.stem,
        "importedAt": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
        "features": features,
        "overlays": [],
        "bundled": True,
    }


def merge_layer(layers, structure, layer, selected_features=None):
    if not layer:
        return
    features = selected_features if selected_features is not None else layer["features"]
    if not features:
        return
    current = layers.setdefault(
        structure,
        {
            "fileName": layer["fileName"],
            "name": structure,
            "importedAt": layer["importedAt"],
            "features": [],
            "overlays": [],
            "bundled": True,
        },
    )
    existing = {
        (
            feature["type"],
            feature["name"],
            json.dumps(feature["coordinates"], sort_keys=True),
        )
        for feature in current["features"]
    }
    for feature in features:
        key = (
            feature["type"],
            feature["name"],
            json.dumps(feature["coordinates"], sort_keys=True),
        )
        if key not in existing:
            current["features"].append(feature)
            existing.add(key)
    current["fileName"] = f"{current['fileName']} + {layer['fileName']}"
    current["importedAt"] = max(current["importedAt"], layer["importedAt"])


def read_default_layers():
    layers = {}
    for path in CORPORATE_ROOT.rglob("*.kmz"):
        normalized = str(path).lower()
        structure = None
        if "pde engenho seco i" in normalized:
            structure = "PDE 1"
        elif "pde jac" in normalized:
            structure = "PDE Jaco"
        if not structure:
            continue
        layer = read_kmz_layer(path)
        if structure == "PDE 1" and layer:
            layer["features"] = [
                feature
                for feature in layer["features"]
                if str(feature["name"]).upper().startswith(("PZ-", "INA-"))
            ]
        merge_layer(layers, structure, layer)

    downloads = [
        DOWNLOADS_ROOT / "ESTRUTURAS GEOTEC.kmz",
        DOWNLOADS_ROOT / "PDE 1.kmz",
        DOWNLOADS_ROOT / "PDE MANGABA.kmz",
    ]
    structure_names = {
        "PDE JACO": "PDE Jaco",
        "PDE MANGABA": "PDE Mangaba",
        "BARRAGEM B4": "Barragem B4",
        "BARRAGEM B1": "Barragem B1",
        "BARRAGEM B2": "PILHA B2",
        "PDE 1": "PDE 1",
    }
    pde1_instruments = {
        "PZ-01",
        "INA-01",
        "INA-02",
        "PZ-04",
        "INA-07",
        "PZ-08",
        "PZ-05",
        "INA-03",
        "INA-09",
        "PZ-06",
    }
    for path in downloads:
        if not path.exists():
            continue
        layer = read_kmz_layer(path)
        if not layer:
            continue
        for source_name, structure in structure_names.items():
            selected = [
                feature
                for feature in layer["features"]
                if (
                    str(feature["name"])
                    .upper()
                    .replace("Ó", "O")
                    .replace("Ô", "O")
                    == source_name
                )
            ]
            merge_layer(layers, structure, layer, selected)
        selected_instruments = [
            feature
            for feature in layer["features"]
            if str(feature["name"]).upper() in pde1_instruments
        ]
        merge_layer(layers, "PDE 1", layer, selected_instruments)
    return layers


def build_inspection_records():
    return [
        {
            "id": "survey123-pde1-20260330-trinca",
            "structure": "PDE 1",
            "date": "2026-03-30",
            "severity": "critical",
            "title": "Trinca superficial na crista",
            "description": (
                "Trinca superficial registrada no banco onde se encontra o PZ-05; "
                "o relatorio recomenda tratamento para evitar evolucao para escorregamento."
            ),
            "instrumentCode": "PZ-05",
            "location": {"latitude": -20.091269, "longitude": -44.11038},
            "source": "Survey123 - FIR PDE I de 30/03/2026",
        },
        {
            "id": "survey123-pde1-20260330-drenagem",
            "structure": "PDE 1",
            "date": "2026-03-30",
            "severity": "warning",
            "title": "Greide e drenagem superficial",
            "description": (
                "Necessidade de intervencao corretiva no greide dos acessos e no "
                "direcionamento da drenagem superficial em pontos localizados."
            ),
            "source": "Survey123 - FIR PDE I de 30/03/2026",
        },
        {
            "id": "survey123-jaco-20260414-drenagem",
            "structure": "PDE Jaco",
            "date": "2026-04-14",
            "severity": "warning",
            "title": "Drenagem com visibilidade reduzida",
            "description": (
                "Vegetacao elevada no entorno dos canais impede a visualizacao "
                "interna e requer intervencao pontual de manutencao."
            ),
            "source": "Survey123 - FIR PDE Jaco de 14/04/2026",
        },
        {
            "id": "survey123-mangaba-20260331-drenagem",
            "structure": "PDE Mangaba",
            "date": "2026-03-31",
            "severity": "warning",
            "title": "Reconformacao de greide",
            "description": (
                "Ponto de atencao para reconformacao do greide e adequacao do "
                "direcionamento da drenagem superficial."
            ),
            "location": {"latitude": -20.089759, "longitude": -44.092416},
            "source": "Survey123 - FIR Mangaba de 31/03/2026",
        },
    ]


def main():
    metrics_file = find_file("Indicadores_Pilhas.xlsx", "18) Indicadores Pilhas")
    rainfall_file = find_file("PLUVIOMETRIA.xlsx", "2) Pluviometria")
    pile_metrics, structure_coordinates = read_pile_metrics(metrics_file)
    rainfall = read_rainfall(rainfall_file)
    payload = {
        "version": "2026-06-12-geoview-operational",
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sources": {
            "pileMetrics": str(metrics_file),
            "rainfall": str(rainfall_file),
        },
        "pileMetrics": pile_metrics,
        "structureCoordinates": structure_coordinates,
        "rainfall": rainfall,
        "rainfallStations": {
            "PDE 1": "PILHA B2",
            "PDE Jaco": "BARRAGEM B4",
            "PDE Mangaba": "BARRAGEM B4",
            "PILHA B2": "PILHA B2",
            "Barragem B1": "BARRAGEM B1",
            "Barragem B4": "BARRAGEM B4",
        },
        "defaultLayers": read_default_layers(),
        "inspections": build_inspection_records(),
    }
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(
        "window.MDSYNC_GEOVIEW_OPERATIONAL = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    print(
        f"Gerado {OUTPUT_FILE} com {len(pile_metrics)} indicadores, "
        f"{len(rainfall)} leituras de chuva e {len(payload['defaultLayers'])} camadas."
    )


if __name__ == "__main__":
    main()
